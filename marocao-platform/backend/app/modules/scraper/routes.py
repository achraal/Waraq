# backend/app/modules/scraper/routes.py
from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks, Body, WebSocket, WebSocketDisconnect
from sqlalchemy.orm import Session
from typing import List
from backend.app.database.connection import get_db
from backend.app.database.models import Tender, EmailNotification
from backend.app.modules.scraper.utils import sync_local_tenders_to_db, export_all_tenders_to_excel
from openpyxl import load_workbook
import os, threading, asyncio
from backend.app.modules.scraper.portal_scraper import run_scraper 
from backend.app.modules.scraper.email_monitor import fetch_marche_publics, save_emails_to_db
from backend.app.database.connection import SessionLocal

# État global pour suivre le scraper (optionnel, pour éviter les doublons)
scraper_status = {"is_running": False}

router = APIRouter(prefix="/scraper", tags=["Scraper & Data Sync"])
DATA_STORAGE_DIR = r"C:\Users\achra\Desktop\Intern\Project\marocao-platform\data_storage"
EXCEL_PATH = os.path.join(DATA_STORAGE_DIR, "tenders_export.xlsx")

@router.post("/sync")
def trigger_sync(db: Session = Depends(get_db)):
    results = sync_local_tenders_to_db(db)
    inserted = results.get("inserted", 0)
    msg = f"Synchronisation terminée. {inserted} nouvelles offres ajoutées." if inserted > 0 else "Base de données déjà à jour."
    return {"status": "success", "message": msg, "details": results}

@router.post("/export-excel")
def trigger_excel_export(db: Session = Depends(get_db)):
    try:
        # 1. Compter les offres en base
        count_db = db.query(Tender).count()
        
        # 2. Vérifier si le fichier existe et compter les lignes (hors header)
        count_excel = 0
        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH, read_only=True)
            ws = wb.active
            # ws.max_row compte toutes les lignes, on enlève 1 pour l'en-tête
            if ws.max_row > 1:
                count_excel = ws.max_row - 1
            wb.close()

        # 3. Comparer et décider
        if count_db > 0 and count_db == count_excel:
            return {
                "status": "success", 
                "message": "Le fichier Excel est déjà à jour (toutes les offres sont présentes).",
                "count": count_db
            }
        
        # 4. Sinon, on effectue l'export
        export_all_tenders_to_excel(db)
        
        return {
            "status": "success", 
            "message": f"Export Excel généré avec succès. {count_db} offres ont été synchronisées.",
            "count": count_db
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'export : {str(e)}")

def run_scraper_task():
    scraper_status["is_running"] = True
    try:
        run_scraper()
    finally:
        scraper_status["is_running"] = False

@router.post("/start-scraping")
def start_scraping(background_tasks: BackgroundTasks):
    if scraper_status["is_running"]:
        raise HTTPException(status_code=400, detail="Le scraper est déjà en cours d'exécution.")
    
    # Lancement en arrière-plan via FastAPI
    background_tasks.add_task(run_scraper_task)
    log_to_frontend("[API] Scraping simple démarré en arrière-plan.")
    return {"status": "success", "message": "Scraping démarré en arrière-plan."}

@router.get("/scraper-status")
def get_status():
    return scraper_status

@router.post("/run-pipeline")
def run_full_pipeline(background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    if scraper_status["is_running"]:
        raise HTTPException(status_code=400, detail="Une tâche est déjà en cours.")

    def execute_pipeline():
        scraper_status["is_running"] = True
        # On crée une session propre à ce thread d'arrière-plan
        db_bg = SessionLocal()
        try:
            log_to_frontend("--- Étape 1 : Démarrage du Scraper ---")
            run_scraper()  # ⚠️ Pense à mettre des log_to_frontend() à l'intérieur de cette fonction aussi !
            
            log_to_frontend("--- Étape 2 : Synchronisation des données ---")
            sync_local_tenders_to_db(db_bg)
            
            log_to_frontend("--- Étape 3 : Export Excel ---")
            export_all_tenders_to_excel(db_bg)
            
            log_to_frontend("--- Pipeline terminé avec succès ---")
        except Exception as e:
            log_to_frontend(f"--- Erreur dans le pipeline : {e} ---")
        finally:
            db_bg.close() # Important : On ferme la session manuellement ici
            scraper_status["is_running"] = False
            log_to_frontend("--- Fin du processus. Scraper arrêté. ---")

    background_tasks.add_task(execute_pipeline)
    return {"status": "success", "message": "Le pipeline complet a été lancé en arrière-plan."}

@router.post("/refresh-emails")
def refresh_emails(db: Session = Depends(get_db)):
    # 1. Récupération depuis le serveur mail
    raw_emails = fetch_marche_publics()
    
    # 2. Sauvegarde dans ta base SQL
    inserted_count = save_emails_to_db(db, raw_emails)
    
    print(f"Sync terminée : {inserted_count} nouveaux messages ajoutés.")
    return {"status": "success", "added": inserted_count}

@router.get("/notifications")
def get_notifications(db: Session = Depends(get_db)):
    # 1. On récupère la liste triée
    notifications = db.query(EmailNotification).order_by(
        EmailNotification.received_at.desc()
    ).all()
    
    # 2. On retourne un dictionnaire avec le compteur et la liste
    return {
        "total": len(notifications),
        "data": notifications
    }

# 1. Marquer une sélection (ou un seul) comme lu
@router.post("/notifications/mark-read")
def mark_notifications_read(
    ids: List[int] = Body(...), 
    db: Session = Depends(get_db)
):
    """Reçoit une liste d'IDs et les marque comme lus."""
    db.query(EmailNotification).filter(EmailNotification.id.in_(ids)).update(
        {"is_read": True}, synchronize_session=False
    )
    db.commit()
    return {"status": "success", "updated_count": len(ids)}

# 2. Marquer TOUT comme lu
@router.post("/notifications/mark-all-read")
def mark_all_read(db: Session = Depends(get_db)):
    """Marque toutes les notifications non lues comme lues."""
    db.query(EmailNotification).filter(EmailNotification.is_read == False).update(
        {"is_read": True}, synchronize_session=False
    )
    db.commit()
    return {"status": "success"}

@router.get("/notifications/unread")
def get_unread_notifications(db: Session = Depends(get_db)):
    """Récupère uniquement les notifications non lues."""
    unread_notifications = db.query(EmailNotification)\
        .filter(EmailNotification.is_read == False)\
        .order_by(EmailNotification.received_at.desc())\
        .all()
        
    return {
        "total": len(unread_notifications),
        "data": unread_notifications
    }

@router.get("/notifications/read")
def get_read_notifications(db: Session = Depends(get_db)):
    """Récupère uniquement les notifications déjà lues."""
    read_notifications = db.query(EmailNotification)\
        .filter(EmailNotification.is_read == True)\
        .order_by(EmailNotification.received_at.desc())\
        .all()
        
    return {
        "total": len(read_notifications),
        "data": read_notifications
    }

# 1. Gestionnaire de WebSockets pour les logs
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            try:
                await connection.send_text(message)
            except:
                pass

manager = ConnectionManager()

# Fonction pour logger à la fois dans le terminal et sur le frontend

def log_to_frontend(message: str):
    print(message) # Affiche dans le terminal Python
    # Envoie au frontend via WebSocket
    try:
        loop = asyncio.get_running_loop()
        loop.create_task(manager.broadcast(message))
    except RuntimeError:
        asyncio.run(manager.broadcast(message))

# 2. Endpoint WebSocket
@router.websocket("/ws/logs")
async def websocket_logs(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # Maintient la connexion ouverte
            data = await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)