from sqlalchemy.orm import joinedload
from backend.app import database
import os, json
from datetime import datetime
from sqlalchemy.orm import Session
from backend.app.database.models import Tender, TenderDocument, TenderLot
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side 
import pandas as pd
import unicodedata

DATA_STORAGE_DIR = r"C:\Users\achra\Desktop\Intern\Project\marocao-platform\data_storage"
EXCEL_PATH = os.path.join(DATA_STORAGE_DIR, "tenders_export.xlsx")

class TenderManager:
    def __init__(self, db_session, data_dir):
        self.db = db_session
        self.data_dir = data_dir

    def is_already_processed(self, ref: str) -> bool:  
        # 1. Vérification BDD (Rien à changer ici, c'est déjà correct)
        exists_in_db = self.db.query(Tender).filter(Tender.reference == ref).first() is not None
        if exists_in_db:
            return True
            
        # 2. Vérification Locale (Corrigée)
        metadata_root = os.path.join(self.data_dir, "metadata")
        
        # On parcourt tous les dossiers de manière récursive
        for root, dirs, files in os.walk(metadata_root):
            for d in dirs:
                # On vérifie si la référence est AU DÉBUT du nom du dossier
                # On utilise .startswith() car c'est plus sûr
                if d.startswith(ref): 
                    return True
        return False

def extract_date_from_path(root_path: str):
    """
    Extrait la date et heure d'extraction à partir de l'arborescence.
    Ex: .../metadata/2026/07/01/04-2026-ENFI_11-07-57
    """
    try:
        # On récupère les parties du chemin
        parts = root_path.split(os.sep)
        # On suppose que les 3 derniers sont AAAA, MM, DD
        day = parts[-2]
        month = parts[-3]
        year = parts[-4]
        
        # Le nom du dossier contient HH-MM-SS
        folder_name = parts[-1]
        time_part = folder_name.split("_")[-1] # Récupère le HH-MM-SS
        
        date_str = f"{year}-{month}-{day} {time_part.replace('-', ':')}"
        return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
    except:
        return datetime.now() # Fallback si erreur

def get_storage_paths(tender_reference: str):
    """
    Crée une structure propre avec un identifiant de microsecondes (batch de temps)
    pour différencier deux exécutions dans la même seconde.
    Structure : data_storage/TYPE/AAAA/MM/DD/REF_HH-MM-SS_batch_XXXXXX/
    """
    now = datetime.now()
    safe_ref = tender_reference.replace("/", "-").replace(" ", "_").strip()
    folder_name = f"{safe_ref}_{now.strftime('%H-%M-%S')}"
    date_path = now.strftime("%Y/%m/%d")
    
    paths = {
        "archive": os.path.join(DATA_STORAGE_DIR, "archives", date_path, folder_name),
        "extracted": os.path.join(DATA_STORAGE_DIR, "extracted", date_path, folder_name),
        "metadata": os.path.join(DATA_STORAGE_DIR, "metadata", date_path, folder_name)
    }
    
    for path in paths.values():
        os.makedirs(path, exist_ok=True)
        
    return paths

def normalize_text(text: str) -> str:
    """Enlève les accents et met en minuscules."""
    nfkd_form = unicodedata.normalize('NFKD', text)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)]).lower()

def get_file_type(filename: str) -> str:
    # On convertit tout en minuscules pour ne plus se soucier des majuscules
    f_norm = normalize_text(filename)
    # f_lower = filename.lower()
    
    # 1. Traitement des cas spécifiques (du plus précis au plus général)
    if f_norm.startswith("avis fr"): return "AVIS_FRANCAIS"
    if f_norm.startswith("avis ar"): return "AVIS_ARABE"
    if f_norm.startswith("avis"): return "AVIS"
    if f_norm.startswith("cps"): return "CPS"
    if f_norm.startswith("rc") or f_norm.startswith("reglement"): return "RC"
    if f_norm.startswith("acte d'engagement"): return "ACTE_ENGAGEMENT"
    if f_norm.startswith("declaration sur l'honneur"): return "DECLARATION_HONNEUR"
    if f_norm.startswith("bdr") or f_norm.startswith("bordereau"): return "BORDEREAU_PRIX"
    
    # 2. Valeur par défaut : si rien ne correspond, on retourne le nom du fichier
    # Exemple: "Plan_architecte.pdf" -> retournera "PLAN_ARCHITECTE.PDF"
    return filename.upper()

def sync_local_tenders_to_db(db: Session):
    """
    Synchronise les métadonnées et documents stockés localement vers la base de données.
    Parcourt récursivement les dossiers pour trouver metadata.json.
    """
    metadata_root = os.path.join(DATA_STORAGE_DIR, "metadata")
    extracted_root = os.path.join(DATA_STORAGE_DIR, "extracted")
    archives_dir = os.path.join(DATA_STORAGE_DIR, "archives")
    
    if not os.path.exists(metadata_root):
        print("-> [Sync] Aucun dossier metadata trouvé.")
        return {"processed": 0, "inserted": 0, "documents_inserted": 0}

    # --- OPTIMISATION : Indexation des chemins d'archives ---
    # On stocke { "nom_dossier": "chemin_complet" } 
    # Cela évite de parcourir l'arborescence à chaque fois
    archives_map = {}
    if os.path.exists(archives_dir):
        for root_arch, _, files_arch in os.walk(archives_dir):
            for file in files_arch:
                if file.endswith(".zip"):
                    # On garde la structure : le nom du dossier sert de clé pour la recherche
                    archives_map[os.path.basename(root_arch)] = os.path.join(root_arch, file).replace("\\", "/")

    inserted_count = 0
    processed_count = 0
    total_docs_count = 0

    registry = {}
    

    # 1. Utilisation de os.walk pour traverser tous les sous-dossiers (AAAA/MM/DD/...)
    for root, dirs, files in os.walk(metadata_root):
        if "metadata.json" in files:
            processed_count += 1
            json_file_path = os.path.join(root, "metadata.json")
            extraction_dt = extract_date_from_path(root)
                
            try:
                with open(json_file_path, 'r', encoding='utf-8') as f:
                    meta_data = json.load(f)
                
                ref = meta_data.get("reference")
                if not ref:
                    continue

                # 2. Vérification stricte des doublons en base de données
                exists = db.query(Tender).filter(Tender.reference == ref).first()
                if exists:
                    continue

                # 1. Chercher le fichier ZIP dynamiquement en parcourant les sous-dossier
                # 1. Préparation de la référence sécurisée pour comparaison
                ref_to_find = ref.replace("/", "-").replace(" ", "_")

                local_zip_path = None

                # --- RECHERCHE OPTIMISÉE (gardant votre logique sémantique) ---
                # On parcourt nos dossiers indexés au lieu de refaire un os.walk()
                for folder_name, zip_full_path in archives_map.items():
                    if ref_to_find in folder_name:
                        local_zip_path = zip_full_path
                        break 
                # -------------------------------------------------------------
                # 1. Récupération des lots
                lots_data = meta_data.get("lots", [])

                # 2. Logique de nettoyage : si on a des lots, on ignore les champs globaux pollués
                # Tu peux définir une liste des champs qui sont souvent mal parsés en cas d'allotissement
                if len(lots_data) > 0:
                    meta_data["caution"] = None
                    meta_data["qualifications"] = None
                    meta_data["agrements"] = None
                    meta_data["prospectus_notices"] = None
                    meta_data["variante"] = None
                    meta_data["provisional_caution"] = None
                    meta_data["reunion"] = None
                    meta_data["visite_lieux"] = None

                             
                # 5. Association des documents extraits
                # Le dossier correspondant dans 'extracted' a le même nom que le dossier contenant metadata.json
                folder_name = os.path.basename(root)
                date_path = os.path.relpath(root, metadata_root) # ex: 2026/06/30/REF_...
                date_dir = os.path.dirname(date_path)
                specific_extracted_dir = os.path.join(extracted_root, date_dir, folder_name)

                # tender_docs_count = 0
                temp_docs = []
                has_nested_zip = False

                if os.path.exists(specific_extracted_dir):  
                    for root_doc, _, files_doc in os.walk(specific_extracted_dir):
                        for filename in files_doc:
                            if filename.lower().endswith('.zip'):
                                has_nested_zip = True
                                continue # On ignore les fichiers ZIP dans la base de documents
                            full_file_path = os.path.join(root_doc, filename)
                            file_type = get_file_type(filename)

                            new_doc = TenderDocument(
                                file_name=filename,
                                file_type=file_type,
                                file_path=full_file_path.replace("\\", "/"),
                            )
                            temp_docs.append(new_doc)
                    #         new_tender.documents.append(new_doc) 
                    #         tender_docs_count += 1 
                    #         total_docs_count += 1 

                    # new_tender.nbr_documents = tender_docs_count


                # 3. Création de l'objet Tender
                if ref not in registry:
                    new_tender = Tender(
                        reference=ref,
                        title=meta_data.get("objet") or "Sans objet",
                        buyer=meta_data.get("acheteur") or "Inconnu",
                        type_annonce=meta_data.get("type_annonce"),
                        procedure=meta_data.get("procedure"),
                        categorie=meta_data.get("categorie"),
                        allotissement=meta_data.get("allotissement"),
                        lieu_execution=meta_data.get("lieu_execution"),
                        estimated_budget=meta_data.get("budget"),
                        reserve_pme=meta_data.get("reserve_pme"),
                        domaines_activite=meta_data.get("domaines_activite"),
                        adresse_retrait=meta_data.get("adresse_retrait"),
                        adresse_depot=meta_data.get("adresse_depot"),
                        lieu_ouverture=meta_data.get("lieu_ouverture"),
                        prix_acquisition=meta_data.get("prix_acquisition"),
                        provisional_caution=meta_data.get("caution"),
                        qualifications=meta_data.get("qualifications"),
                        agrements=meta_data.get("agrements"),
                        variante=meta_data.get("variante"),
                        deadline=meta_data.get("deadline"),
                        prospectus_notices=meta_data.get("prospectus_notices"),
                        reunion=meta_data.get("reunion"),
                        visite_lieux=meta_data.get("visite_lieux"),
                        contact_administratif=meta_data.get("contact_administratif"),
                        local_zip_path=local_zip_path,
                        extraction_date=extraction_dt,
                        #nbr_lots=len(lots_data),
                        nbr_lots=len(lots_data) if len(lots_data) > 0 else 1,
                        nbr_documents=len(temp_docs),
                        is_recursive=meta_data.get('is_recursive', has_nested_zip),
                        metadata_json=meta_data,
                    )
                    # 4. Association des LOTS (Nouveau !)
                    lots_data = meta_data.get("lots", []) # On récupère la liste des lots
                    for lot_info in lots_data:
                        new_lot = TenderLot(
                            #tender_id=new_tender.id,
                            lot_number=lot_info.get("lot_number"),
                            title=lot_info.get("title"),
                            description=lot_info.get("description"),
                            estimated_budget=lot_info.get("estimated_budget"),
                            provisional_caution=lot_info.get("provisional_caution"),
                            qualifications=lot_info.get("qualifications"),
                            agrements=lot_info.get("agrements"),
                            prospectus_notices=lot_info.get("prospectus_notices"),
                            reunion=lot_info.get("reunion"),
                            visite_lieux=lot_info.get("visite_lieux"),
                            variante=lot_info.get("variante"),
                            env_considerations=lot_info.get("env_considerations"),
                            reserve_pme=lot_info.get("reserve_pme")
                        )
                        new_tender.lots.append(new_lot)
                    registry[ref] = {"tender": new_tender, "docs": []}
                else:
                    # Si on l'a déjà vu, on récupère l'instance existante pour y ajouter des docs
                    new_tender = registry[ref]["tender"]
                    if has_nested_zip:
                        new_tender.is_recursive = True

                # 5. Association des documents extraits
                for doc in temp_docs:
                    new_tender.documents.append(doc)
                    total_docs_count += 1

                new_tender.nbr_documents = len(new_tender.documents)

                #db.add(new_tender)
                #inserted_count += 1
            except Exception as e:
                print(f"Erreur lors du traitement de {root} : {e}")
                db.rollback()

    # 2. Insertion finale fusionnée dans la BDD
    for ref, data in registry.items():
        tender = data["tender"]
        
        # Vérification finale au cas où une sync précédente a laissé des traces
        exists = db.query(Tender).filter(Tender.reference == ref).first()
        if not exists:
            db.add(tender)
            inserted_count += 1
    
    if inserted_count > 0:
        db.commit()
        
    print(f"-> [Sync BDD] Fini. {inserted_count} offres et {total_docs_count} documents ajoutés.")
    return {"processed": processed_count, "inserted": inserted_count, "documents_inserted": total_docs_count}


def append_to_excel_fast(meta_data: dict, sync_date: str, extraction_date: datetime):
    """Ajoute une ligne de manière optimisée avec openpyxl."""
    print(f"-> [Excel] Tentative d'ajout pour la référence : {meta_data.get('reference')}")
    print(f"-> [Excel] Chemin du fichier : {EXCEL_PATH}")
    # Liste des colonnes dans l'ordre pour garantir la cohérence  
    headers = [
        "Date d'importation", "Date d'extraction", "Référence", "Objet", "Acheteur public", 
        "Type d'annonce", "Procédure", "Catégorie principale", "Allotissement", 
        "Lieu d'exécution", "Estimation (en Dhs TTC)", "Réservé à la TPE et PME installées au Maroc", 
        "Domaines d'activité", "Adresse de retrait", "Adresse de dépôt", 
        "Lieu d'ouverture", "Prix d'acquisition", "Caution provisoire", 
        "Qualifications", "Agréments", "Variante", 
        "Date et heure limite de remise des plis", "Prospectus, notices", 
        "Réunion", "Visites des lieux", "Contact Administratif"
    ]

    # Préparation des données dans l'ordre des headers
    row_data = [
        sync_date, extraction_date.strftime("%Y-%m-%d %H:%M:%S"),meta_data.get("reference"), meta_data.get("objet"), meta_data.get("acheteur"),
        meta_data.get("type_annonce"), meta_data.get("procedure"), meta_data.get("categorie"),
        meta_data.get("allotissement"), meta_data.get("lieu_execution"), meta_data.get("budget"),
        meta_data.get("reserve_pme"), meta_data.get("domaines_activite"), meta_data.get("adresse_retrait"),
        meta_data.get("adresse_depot"), meta_data.get("lieu_ouverture"), meta_data.get("prix_acquisition"),
        meta_data.get("caution"), meta_data.get("qualifications"), meta_data.get("agrements"),
        meta_data.get("variante"), meta_data.get("deadline"), meta_data.get("prospectus_notices"),
        meta_data.get("reunion"), meta_data.get("visite_lieux"), meta_data.get("contact_administratif")
    ]

    # Si le fichier n'existe pas, on le crée avec les en-têtes
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(EXCEL_PATH)
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        ws.append(row_data)
        apply_professional_style(ws)
        wb.save(EXCEL_PATH) 
    except PermissionError:
        print("!!! ERREUR : Fermez le fichier Excel pour permettre la mise à jour.")
 

def export_all_tenders_to_excel(db: Session):
    print(f"-> [Excel] Export complet de la base vers Excel...")
    
    wb = Workbook()
    ws = wb.active
    
    # 1. Tes headers actuels + colonnes pour les lots
    headers = [
        "Date d'importation", "Date d'extraction", "Référence", "Objet", "Acheteur public", 
        "Type d'annonce", "Procédure", "Catégorie principale", "Allotissement", 
        "Lieu d'exécution", "Estimation (en Dhs TTC)", "Réservé à la TPE et PME", 
        "Domaines d'activité", "Adresse de retrait", "Adresse de dépôt", 
        "Lieu d'ouverture", "Prix d'acquisition", "Caution provisoire", 
        "Qualifications", "Agréments", "Variante", 
        "Date limite", "Prospectus/Notices", "Réunion", "Visites des lieux", 
        "Contact Administratif", "Nombre de lots",
        "Détails des Lots" # <-- La nouvelle colonne synthétique
    ]
    ws.append(headers)
    
    # tenders = db.query(Tender).all()
    tenders = db.query(Tender).options(joinedload(Tender.lots)).all()
    sync_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    for tender in tenders:
        meta = tender.metadata_json or {}
        extraction_dt = tender.extraction_date if tender.extraction_date else datetime.now()
        
        # 2. Construction synthétique des lots (Concaténation)
        # Construction détaillée et exhaustive de chaque lot
        lots_text = ""
        if tender.lots:
            for lot in tender.lots:
                lots_text += (
                    f"Lot N°{lot.lot_number} : {lot.title}\n"
                    f"   - Description : {lot.description or '-'}\n"
                    f"   - Budget : {lot.estimated_budget or '-'}\n"
                    f"   - Caution : {lot.provisional_caution or '-'}\n"
                    f"   - Qualifications : {lot.qualifications or '-'}\n"
                    f"   - Agréments : {lot.agrements or '-'}\n"
                    f"   - Prospectus/Notices : {lot.prospectus_notices or '-'}\n"
                    f"   - Réunion : {lot.reunion or '-'}\n"
                    f"   - Visite des lieux : {lot.visite_lieux or '-'}\n"
                    f"   - Variante : {lot.variante or '-'}\n"
                    f"   - Considérations Env. : {lot.env_considerations or '-'}\n"
                    f"   - Réserve PME : {lot.reserve_pme or '-'}\n"
                    f"--------------------------------------------\n"
                )
        else:
            lots_text = "Aucun lot spécifique"
        
        # 3. Ligne de données complète
        row_data = [
            sync_date, 
            extraction_dt.strftime("%Y-%m-%d %H:%M:%S"),
            meta.get("reference"), meta.get("objet"), meta.get("acheteur"),
            meta.get("type_annonce"), meta.get("procedure"), meta.get("categorie"),
            meta.get("allotissement"), meta.get("lieu_execution"), meta.get("budget"),
            meta.get("reserve_pme"), meta.get("domaines_activite"), meta.get("adresse_retrait"),
            meta.get("adresse_depot"), meta.get("lieu_ouverture"), meta.get("prix_acquisition"),
            meta.get("caution"), meta.get("qualifications"), meta.get("agrements"),
            meta.get("variante"), meta.get("deadline"), meta.get("prospectus_notices"),
            meta.get("reunion"), meta.get("visite_lieux"), meta.get("contact_administratif"), tender.nbr_lots,
            lots_text # <-- Insertion des lots ici
        ]
        ws.append(row_data) 
    
    # 4. Finalisation
    apply_professional_style(ws)
    #ensure_directory_exists(EXCEL_PATH) # S'assure que le dossier existe
    wb.save(EXCEL_PATH)
    print(f"-> [Excel] Export terminé avec succès.") 

def apply_professional_style(ws):
    """Applique un style professionnel, filtres et bordures."""
    
    # Définition des styles
    header_font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 1. Appliquer le style au Header (Ligne 1)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # 2. Activer le filtre automatique (la petite flèche)
    # On définit la zone du filtre sur toute la première ligne
    ws.auto_filter.ref = ws.dimensions

    # 3. Style des données (Lignes 2+)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    # 4. Ajustement automatique de la largeur
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                # On force la conversion en str pour éviter les erreurs
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except:
                pass
        # Largeur minimale de 15, maximale de 50
        ws.column_dimensions[column].width = max(15, min(max_length + 2, 50))