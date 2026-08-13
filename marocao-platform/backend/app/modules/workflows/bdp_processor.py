import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class BDPProcessor:
    """
    Service de traitement des Bordereaux des Prix. Fonctionnalités :
    - Détection automatique du format.
    - Extraction des lignes depuis PDF natif.
    - Extraction des lignes depuis PDF scanné via OCR/ONNX.
    - Extraction depuis XLS/XLSX/XLSM.
    - Extraction depuis DOC/DOCX via le service OCR.
    - Détection plus souple des colonnes.
    - Préparation des champs à remplir par l'utilisateur.
    - Injection des prix unitaires.
    - Calcul HT / TVA / TTC.
    """

    NUMBER_PATTERN = re.compile(r"^\s*\d+(?:[.\-]\d+)*\s*$")
    QUANTITY_PATTERN = re.compile(r"^\s*\d+(?:[.,]\d+)?\s*$")
    UNIT_PATTERN = re.compile(r"^[A-Za-zÀ-ÿ0-9²³/%\-\.]+$")
    HEADER_ALIASES = {
        "number": ["n°","no","numéro","numero","n","article","prix","n prix"],
        "description": ["désignation","designation","description","libellé","libelle","objet",],
        "unit": ["unité","unite","u","unit",],
        "quantity": ["quantité","quantite","qte","qté","quant",],
        "unit_price": ["prix unitaire","pu","p.u","prix unitaire ht",],
        "total": ["montant","total","montant ht","prix total","total ht",],
    }

    # UTILITAIRES

    @staticmethod
    def _normalize(value: Any) -> str:
        """
        Normalise une valeur provenant d'un fichier Excel ou OCR.
        """
        if value is None:
            return ""
        text = str(value).strip().lower()
        replacements = {"é": "e","è": "e","ê": "e","ë": "e","à": "a","â": "a","ä": "a","î": "i","ï": "i","ô": "o","ö": "o","ù": "u","û": "u","ü": "u",}

        for old, new in replacements.items():
            text = text.replace(old, new)
        text = re.sub(r"\s+", " ", text)
        return text

    @staticmethod
    def _parse_number(value: Any) -> Optional[float]:
        """
        Convertit une valeur numérique provenant d'un BDP.
        """
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if not text:
            return None
        text = (text.replace(" ", "").replace("\u00a0", ""))
        # Gestion des nombres du type 1 250,50
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "")
                text = text.replace(",", ".")
            else:
                text = text.replace(",", "")
        else:
            text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    # PARSING TEXTE / OCR
    def parse_bdp_text(self,text_content: str,) -> List[Dict[str, Any]]:
        """
        Extrait les lignes d'un BDP depuis du texte natif ou OCR.
        Cette méthode est utilisée aussi bien pour :
        - PDF natif
        - PDF scanné
        - DOC
        - DOCX
        - image
        Le texte peut provenir de RapidOCR/ONNX.
        """
        items: List[Dict[str, Any]] = []
        if not text_content:
            logger.warning("[BDP] Aucun texte à analyser.")
            return items

        for raw_line in text_content.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            # Nettoyage des espaces OCR
            line = re.sub(r"\s+", " ", line)
            # Recherche du numéro de prix au début.
            match_number = re.match(r"^(\d+(?:[.\-]\d+)*)\s+(.*)$",line,)
            if not match_number:
                continue
            number = match_number.group(1)
            remaining = match_number.group(2).strip()
            tokens = remaining.split()
            if len(tokens) < 3:
                continue
            quantity_index = None
            # On cherche la quantité vers la fin.
            for index in range(len(tokens) - 1, -1, -1):
                if self._parse_number(tokens[index]) is not None:
                    quantity_index = index
                    break
            if quantity_index is None:
                continue
            quantity = self._parse_number(tokens[quantity_index])
            if quantity is None:
                continue
            before_quantity = tokens[:quantity_index]
            if len(before_quantity) < 2:
                continue
            unit = before_quantity[-1]
            if not self.UNIT_PATTERN.match(unit):
                continue
            description = " ".join(before_quantity[:-1]).strip()
            if not description:
                continue

            items.append(
                {
                    "item_number": number,
                    "description": description,
                    "unit": unit,
                    "quantity": quantity,
                    "unit_price": None,
                    "total_price_ht": None,
                }
            )
        logger.info("[BDP] Texte analysé | items=%s",len(items),)
        return items

    # PARSING EXCEL

    def _detect_excel_columns(self,headers: List[str],) -> Dict[str, int]:
        """
        Détecte automatiquement les colonnes d'un BDP Excel.
        Les colonnes ne sont pas supposées être toujours dans le même ordre.
        """
        mapping: Dict[str, int] = {}
        normalized_headers = [self._normalize(header) for header in headers]
        for field, aliases in self.HEADER_ALIASES.items():
            normalized_aliases = [self._normalize(alias) for alias in aliases]
            for index, header in enumerate(normalized_headers):
                if not header:
                    continue
                for alias in normalized_aliases:
                    if header == alias or alias in header:
                        mapping[field] = index
                        break
                if field in mapping:
                    break
        return mapping

    def parse_excel(self,file_path: str | Path,) -> List[Dict[str, Any]]:
        """
        Extrait automatiquement les lignes d'un BDP Excel.
        Supporte :  - XLSX   - XLSM
        Les colonnes peuvent être dans un ordre différent.
        """
        path = Path(file_path)
        items: List[Dict[str, Any]] = []
        workbook = load_workbook(filename=str(path),data_only=False,)
        try:
            for worksheet in workbook.worksheets:
                rows = list(worksheet.iter_rows(values_only=True))
                if not rows:
                    continue
                # Recherche de la ligne d'en-tête.
                header_row_index = None
                column_mapping = {}

                for index, row in enumerate(rows[:20]):
                    headers = [str(value).strip() if value is not None else "" for value in row]
                    mapping = self._detect_excel_columns(headers)
                    if "number" in mapping and "description" in mapping:
                        header_row_index = index
                        column_mapping = mapping
                        break

                # Cas où aucun header n'est détecté. On utilise la structure classique :
                # N° | Désignation | Unité | Quantité
                if header_row_index is None:
                    logger.warning(
                        "[BDP] Header non détecté | "
                        "sheet=%s | fallback colonnes 0-3",
                        worksheet.title,
                    )
                    column_mapping = {"number": 0,"description": 1,"unit": 2,"quantity": 3,}
                    header_row_index = -1

                # Lecture des lignes
                for row in rows[header_row_index + 1 :]:
                    if not row:
                        continue
                    def get_value(field: str,) -> Any:
                        index = column_mapping.get(field)
                        if index is None:
                            return None
                        if index >= len(row):
                            return None
                        return row[index]
                    number_raw = get_value("number")
                    if number_raw is None:
                        continue
                    number = str(number_raw).strip()
                    if not self.NUMBER_PATTERN.match(number):
                        continue

                    description = str(get_value("description") or "").strip()
                    unit = str(get_value("unit") or "").strip()
                    quantity = self._parse_number(get_value("quantity"))
                    unit_price = self._parse_number(get_value("unit_price"))
                    total_price = self._parse_number(get_value("total"))

                    if not description:
                        continue
                    items.append(
                        {
                            "item_number": number,
                            "description": description,
                            "unit": unit,
                            "quantity": quantity,
                            "unit_price": unit_price,
                            "total_price_ht": total_price,
                        }
                    )
        finally:
            workbook.close()

        logger.info("[BDP] Excel analysé | file=%s | items=%s",path,len(items),)
        return items

    # EXTRACTION PRINCIPALE
    def extract_structure(self,file_path: str | Path,ocr_service=None,) -> Dict[str, Any]:
        """
        Analyse automatiquement un BDP quel que soit son format.
        PDF : - texte natif -> extraction native - scan -> OCR ONNX
        XLS/XLSX/XLSM : - extraction directe des cellules
        DOC/DOCX/images : - extraction via OCR service.
        Retourne également les métadonnées permettant de savoir
        si le document était scanné.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"BDP introuvable : {path}")
        suffix = path.suffix.lower()
        logger.info("[BDP] Début analyse | file=%s | format=%s",path,suffix,)
        result: Dict[str, Any] = {
            "file_path": str(path),
            "format": suffix,
            "is_scanned": False,
            "inspection_method": None,
            "items": [],
            "items_count": 0,
        }
        # Excel
        if suffix in {".xlsx","xlsm"}:
            items = self.parse_excel(path)
            result.update({"items": items,"items_count": len(items),"inspection_method":"NATIVE_EXCEL",})

        # PDF / DOC / DOCX / Images
        else:
            if ocr_service is None:
                raise ValueError("OCR requis pour analyser ce BDP.")
            ocr_result = ocr_service.extract(str(path))
            text = ocr_result.get("text","")
            items = self.parse_bdp_text(text)
            result.update(
                {
                    "items": items,
                    "items_count": len(items),
                    "is_scanned": bool(ocr_result.get("is_scanned",False)),
                    "inspection_method": ocr_result.get("inspection_method"),
                    "page_count": ocr_result.get("page_count"),
                    "ocr_duration_sec": ocr_result.get("ocr_duration_sec",0),
                }
            )

        # Vérification
        if not result["items"]:
            logger.warning("[BDP] Aucun article détecté | " "file=%s | method=%s",
                path, result.get("inspection_method",),
            )
        else:
            logger.info(
                "[BDP] Analyse terminée | "
                "file=%s | items=%s | scanned=%s | method=%s",
                path,result["items_count"],result["is_scanned"],result["inspection_method"],
            )
        return result

    # CHAMPS À REMPLIR
    def get_fields_to_fill(self, items: List[Dict[str, Any]] ) -> List[Dict[str, Any]]:
        """
        Prépare les champs que l'utilisateur doit remplir.
        Le prix unitaire est demandé pour chaque ligne qui ne possède pas déjà de valeur.
        """
        fields = []
        for item in items:
            fields.append(
                {
                    "item_number": item["item_number"],
                    "description": item["description"],
                    "unit": item["unit"],
                    "quantity": item["quantity"],
                    "field": "unit_price",
                    "label": f"Prix unitaire HT ({item['unit']})",
                    "value": item.get("unit_price"),
                    "required": True,
                }
            )

        logger.info("[BDP] Champs à remplir préparés | count=%s", len(fields),)
        return fields

    # CALCULS

    def calculate_totals(self, items: List[Dict[str, Any]], tva_rate: float = 0.20,) -> Dict[str, Any]:
        """
        Calcule les montants HT, TVA et TTC du BDP.
        """
        total_ht = 0.0
        for item in items:
            quantity = (item.get("quantity") or 0)
            unit_price = (item.get("unit_price") or 0)
            total = (quantity * unit_price)
            item["total_price_ht"] = round(total,2,)
            total_ht += total
        total_tva = (total_ht * tva_rate)
        total_ttc = (total_ht + total_tva)
        totals = {"total_ht": round(total_ht,2,),
            "tva_amount": round(total_tva,2,),
            "total_ttc": round(total_ttc,2,),
            "items_count": len(items),
        }
        logger.info(
            "[BDP] Totaux calculés | "
            "HT=%s | TVA=%s | TTC=%s",
            totals["total_ht"], totals["tva_amount"], totals["total_ttc"],
        )
        return totals

    # REMPLISSAGE
    def fill_items(self,items: List[Dict[str, Any]],values: List[Dict[str, Any]],) -> Dict[str, Any]:
        """
        Injecte les prix unitaires fournis par l'utilisateur.
        Les montants de chaque ligne et les totaux sont ensuite recalculés.
        """

        values_map = { str(item["item_number"]): item for item in values }
        for item in items:
            key = str(item["item_number"])
            submitted = values_map.get(key)
            if not submitted:
                continue
            unit_price = self._parse_number(submitted.get("unit_price"))
            item["unit_price"] = unit_price

        totals = self.calculate_totals(items)

        logger.info(
            "[BDP] BDP complété | "
            "items=%s | total_ht=%s",
            len(items),
            totals["total_ht"],
        )
        return {"items": items, "summary": totals,}